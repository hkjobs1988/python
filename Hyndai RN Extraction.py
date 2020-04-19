from selenium import webdriver
import getpass
import re
import shutil
from tkinter import messagebox
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver import ActionChains
from selenium.webdriver.support.ui import Select
import openpyxl
import win32com.client as win32
import time
import datetime
from selenium.webdriver.common.keys import Keys
import psutil
from selenium.common.exceptions import TimeoutException
# from datetime import datetime, time as dt_time
import PyPDF2
import winsound
import os
import glob

for proc in psutil.process_iter():
    proc_name = proc.name()
    if proc_name == 'chrome.exe':
        print('Closing the Running Chrome.')
        proc.kill()
        time.sleep(1)
    else:
        pass

wait=60
system_user = getpass.getuser()
if not os.path.exists('C:\HYUNDAI\PDF'):
    os.makedirs('C:\HYUNDAI\PDF')

daily_file_path = 'C:\\HYUNDAI\\DATA.xlsx'
daily_file_path_backup = 'C:\\HYUNDAI\\DATA_up.xlsx'

daily_workbook = openpyxl.load_workbook(daily_file_path)
data_Sheet_daily_workbook = daily_workbook["Data"]
details_sheet = daily_workbook["Details"]

daily_workbook.save(daily_file_path)
daily_workbook.save(daily_file_path)

# Login Portal

user_id=details_sheet.cell(row=1, column=2).value
password=details_sheet.cell(row=2, column=2).value
emails = details_sheet.cell(row=3, column=2).value


# Check if Element is available else refresh the page and try again
login_attempt_count = 1
while True:
    try:
        opt = webdriver.ChromeOptions()
        prefs = {"plugins.always_open_pdf_externally": True}

        opt.add_experimental_option("prefs", prefs)

        opt.add_argument("--ignore-certificate-errors")
        opt.add_argument("--start-maximized")
        driver = webdriver.Chrome(executable_path="C:\\chrome driver\\chromedriver.exe", options=opt)
        driver.implicitly_wait(10)
        driver.get('https://www.hyundaiassurance.in/LV/login.aspx')
        try:
            time.sleep(5)
            alert = driver.switch_to.alert
            alert.accept()
        except Exception as e:
            pass
        time.sleep(2)
        WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.ID, 'btnLogin')))
        Flag_Login_Option = 'Found'
        print ('Login Button  Found')

    except TimeoutException:
        print ('Login Button Not Found')
        Flag_Login_Option = 'Not Found'
        login_attempt_count=login_attempt_count+1
        print (login_attempt_count)
        driver.quit()
        time.sleep(4)
    if Flag_Login_Option=='Found':
        break
    elif login_attempt_count>3:
        driver.quit()
        print ('Not Able To Start')
        exit()

print (' Final out'+Flag_Login_Option)

driver.find_element_by_id('txtUserName').click()
driver.find_element_by_id('txtUserName').clear()
# driver.find_element_by_id('txtUserName').send_keys('Bhraxa01')
driver.find_element_by_id('txtUserName').send_keys(str(user_id))

time.sleep(2)
driver.find_element_by_id('txtPassword').click()
driver.find_element_by_id('txtPassword').clear()
# driver.find_element_by_id('txtPassword').send_keys('Today@123')
driver.find_element_by_id('txtPassword').send_keys(str(password))


driver.find_element_by_id('btnLogin').send_keys(Keys.ENTER)

WebDriverWait(driver,wait).until(EC.element_to_be_clickable((By.LINK_TEXT,'Report'))),time.sleep(2)
driver.find_element_by_xpath('//*[text()=" Report "]').click()
time.sleep(2)
driver.find_element_by_link_text('Renewal Notice').click(), time.sleep(2)

for child_rowno, child_row in enumerate(data_Sheet_daily_workbook.iter_rows(min_col=1, min_row=2, max_row=data_Sheet_daily_workbook.max_row)):
    # Set Variables for Further action on single file
    policy_number = data_Sheet_daily_workbook.cell(row=child_rowno + 2, column=8).value
    print(policy_number)
    expiry_date = data_Sheet_daily_workbook.cell(row=child_rowno + 2, column=7).value
    pdf_policy_number = data_Sheet_daily_workbook.cell(row=child_rowno + 2, column=12).value
    if 'HAX' in str(pdf_policy_number):
        continue
    xpath_rn_policy_number = '//*[@id="ContentPlaceHolder1_txtPolicyNumber"]'
    xpath_from_date = '//*[@id="ContentPlaceHolder1_txtStartDate"]'
    xpath_to_date = '//*[@id="ContentPlaceHolder1_txtEndDate"]'
    # First Click on Reset

    driver.find_element_by_xpath('//*[@id="ContentPlaceHolder1_btnReset"]').click()
    time.sleep(3)
    driver.find_element_by_xpath('//*[@id="ContentPlaceHolder1_divLoading"]/table/tbody/tr/td/img').click()
    WebDriverWait(driver,wait).until(EC.invisibility_of_element((By.XPATH,'//*[@id="ContentPlaceHolder1_divLoading"]/table/tbody/tr/td/img')))
    time.sleep(1)
    WebDriverWait(driver,wait).until(EC.element_to_be_clickable,((By.XPATH,xpath_rn_policy_number)))
    time.sleep(2)
    # Send details
    driver.find_element_by_xpath(xpath_rn_policy_number).click()
    driver.find_element_by_xpath(xpath_rn_policy_number).clear()
    driver.find_element_by_xpath(xpath_rn_policy_number).send_keys(policy_number)
    time.sleep(1)
    driver.find_element_by_xpath(xpath_from_date).click()
    driver.find_element_by_xpath(xpath_from_date).clear()
    driver.find_element_by_xpath(xpath_from_date).send_keys(expiry_date.strftime('%d/%b/%Y'))
    time.sleep(1)
    driver.find_element_by_xpath(xpath_to_date).click()
    driver.find_element_by_xpath(xpath_to_date).clear()
    driver.find_element_by_xpath(xpath_to_date).send_keys(expiry_date.strftime('%d/%b/%Y'))

    # Click Go
    driver.find_element_by_xpath('//*[@id="ContentPlaceHolder1_btnSearch"]').click()
    # Wait for the table to be loaded
    # waith for JS Status whee to go off
    time.sleep(3)

    WebDriverWait(driver,wait*2).until(EC.element_to_be_clickable,((By.XPATH,'//*[@id="ContentPlaceHolder1_grdRenewalLetter_chkbxRenewal_0"]')))
    # Select the check box

    chk_attempt_count=1
    status = False

    while True:
        try:
            print('Waiting for check box....')
            time.sleep(10)
            xpath_check_box = '//*[@id="ContentPlaceHolder1_grdRenewalLetter_chkbxRenewal_0"]'
            driver.find_element_by_xpath(xpath_check_box).click()
            # Click Generate
            driver.find_element_by_xpath('//*[@id="ContentPlaceHolder1_btnGenerate"]').send_keys(Keys.ENTER)
            time.sleep(1)
            if  driver.find_element_by_xpath(xpath_check_box).is_selected():
                status=True
        except Exception as e:
            pass
        chk_attempt_count = chk_attempt_count+1
        if status==True:
            break
        elif chk_attempt_count>=30:
            break
    main_window = driver.current_window_handle
    time.sleep(4)
    handles = driver.window_handles
    print('Moving To Popup')
    time.sleep(3)
    for ii, hh in enumerate(handles):
        if 'https://www.hyundaiassurance.in/LV/Renewal/' in str(driver.current_url):
            print(driver.switch_to.window(hh))


    WebDriverWait(driver,wait*5).until(EC.element_to_be_clickable,((By.XPATH,'//*[@id="rptViewer_ctl06_ctl04_ctl00_ButtonImg"]')))

    time.sleep(1)
    download_icon = driver.find_element_by_xpath('//*[@id="rptViewer_ctl06_ctl04_ctl00_ButtonImg"]')
    ActionChains(driver).move_to_element(download_icon).pause(2).click(download_icon).pause(2).send_keys(Keys.ENTER).perform()

    # Check for the latest file
    while True:
        time.sleep(5)
        print('Checking Time Stamp')
        list_of_files = glob.glob('C:\\Users\\' + system_user + '\\Downloads\\Renewal*pdf')
        latest_file = max(list_of_files, key=os.path.getctime)
        file_age = datetime.datetime.now() - datetime.datetime.fromtimestamp(os.path.getmtime(latest_file))
        print('Latest PDF File ' + latest_file + '-- Age of it ' + str(file_age))
        if file_age < datetime.timedelta(seconds=30):
            break
    print('Final Latest File - ' + latest_file)

    shutil.copy(latest_file,'C:\\HYUNDAI\\PDF\\'+policy_number.replace('HAX/','')+'.pdf')
    print('copied')
    time.sleep(1)
    os.remove(latest_file)
    # Close the Window

    handles = driver.window_handles
    print('Moving To Popup Window to Close')
    try:
        for ii, hh in enumerate(handles):
            print(hh.title)
            # print (driver.current_url)
            if 'https://www.hyundaiassurance.in/LV/Renewal/VSRenewalLetter.aspx?IsAddon' in str(driver.current_url):
                driver.close()
                break
    except Exception as e:
        pass

    # Move to main window
    driver.switch_to.window(main_window)

    daily_workbook.save(daily_file_path)
    daily_workbook.save(daily_file_path)

    # Read Text From PDF*****************************

    pdf_file = open('C:\\HYUNDAI\\PDF\\'+policy_number.replace('HAX/','')+'.pdf', 'rb')
    read_pdf = PyPDF2.PdfFileReader(pdf_file)
    time.sleep(1)
    number_of_pages = read_pdf.getNumPages()
    page = read_pdf.getPage(0)
    page_content = page.extractText()
    # print(str(page_content))
    list_content = str(page_content).split('\n')
    print('*** Converting from PDF to Text ***')
    pdf_policy_number = list_content[9]

    vehical_reg_num = fuel_type = list_content[23]
    print('Before -'+str(vehical_reg_num))
    if not str(vehical_reg_num[-2]).isdigit() :
        vehical_reg_num=None
        list_content.insert(23, None)


    fuel_type = list_content[24]
    print(fuel_type)
    chassis_num = list_content[25] + list_content[26]
    engine_num = list_content[27]
    yom = list_content[28]
    print(yom)
    dept_idv = list_content[35]
    ele_acc_value = list_content[36]
    non_elc_acc_value = list_content[37]
    cng_lpg_value = list_content[38]
    total_idv = list_content[39]
    print('Total IDV - ' + str(total_idv))
    a, premium1 = str(page_content).split('Own Damage Premium (A)\n', 1)
    print('**** premium 1')
    # Basic Premium
    a, basic_premium = str(premium1).split('Basic Premium \n', 1)
    basic_premium, b = str(basic_premium).split('\n', 1)
    # print('*** basic premium')
    # print(basic_premium)

    # NCB %
    # print(premium1)
    a, ncb = str(premium1).split('Less : NCB', 1)
    ncb, b = str(ncb).strip().split('\n', 1)
    print('*** ncb')
    ncb, b = str(ncb).split('%)', 1)
    ncb=str(ncb).replace('(','')
    print(ncb)

    # NCB Value
    # NCB %
    # print(premium1)
    a, ncb_value = str(premium1).split('Less : NCB', 1)
    # print('*** ncb VALUE******************')
    # print(ncb_value)
    c,ncb_value = str(ncb_value).split(' \n',1)
    print('*** new ncb VALUE******************')
    ncb_value,a=str(ncb_value).split('\n',1)
    ncb_value =str(ncb_value).replace('Electrical Accessories','')
    print(ncb_value)
    # Electrical Accessories
    a, ele_accessories = str(premium1).split('Electrical Accessories', 1)
    ele_accessories, b = str(ele_accessories).strip().split('\n', 1)
    # print('*** ELE Access')
    # print(ele_accessories)

    # Other Discount

    a, other_discount = str(premium1).split('Less : Other Discounts', 1)
    other_discount, b = str(other_discount).strip().split('\n', 1)
    # print('*** other discount')
    # print(other_discount)

    # Non Electrical Accessories
    a, non_ele_accessories = str(premium1).split('Non \nElectrical Accessories', 1)
    non_ele_accessories, b = str(non_ele_accessories).strip().split('\n', 1)
    # print('*** Non Electrical')
    # print(non_ele_accessories)

    # Total OD Premium
    a, total_od_premium = str(premium1).split('Total Own Damage Premium', 1)
    total_od_premium, b = str(total_od_premium).strip().split('\n', 1)
    # print('*** od premium')
    # print(total_od_premium)

    # CNG LPG
    a, lpg_cng_premium = str(premium1).split('CNG/LPG Kit', 1)
    lpg_cng_premium, b = str(lpg_cng_premium).strip().split('\n', 1)
    # print('*** lpg cng')
    # print(lpg_cng_premium)

    # IMT 58 Premium

    a, imt58_premium = str(premium1).split('IMT 58 Premium', 1)
    imt58_premium, b = str(imt58_premium).strip().split('\n', 1)
    # print('*** IMT 58')
    # print(imt58_premium)

    # Basic Premium
    a, total_basic_premium = str(premium1).split('Sub Total (Basic Premium)', 1)
    total_basic_premium, b = str(total_basic_premium).strip().split('\n', 1)
    # print('*** basic premium')
    # print(total_basic_premium)

    # Add on Permium
    addon_premium, a = str(premium1).split('Net Own Damage Premium (A)', 1)

    a, addon_premium = str(addon_premium).split('Add On Cover  Premium', 1)
    # addon_premium,b =str(addon_premium).strip().split('\n',1)
    add_on_text = addon_premium

    addon_premium_final = re.search(r'\d+', addon_premium).group()
    print('*** Add on premium*************************************')
    print(addon_premium_final)

    # NET OD premium
    a, net_od_premium = str(premium1).split('Net Own Damage Premium (A)', 1)
    net_od_premium, b = str(net_od_premium).strip().split('\n', 1)
    print('*** NET Premium')
    # print(net_od_premium)

    if 'Liability Premium (B)' in str(premium1):
        # Basic TP premium
        # print(premium1)

        a, basic_tp_premium = str(premium1).split('Basic Third Party Liability',1)
        basic_tp_premium, b = str(basic_tp_premium).strip().split('\n', 1)
        # print('*** BASIC TP')
        # print(basic_tp_premium)

        # PA for Owner Driver
        a, pa_to_owner_driver_premium = str(premium1).split('PA Cover For Owner Driver (IMT-15)', 1)
        pa_to_owner_driver_premium, b = str(pa_to_owner_driver_premium).strip().split('\n', 1)
        # print('*** PA Cover for Paid Driver')
        # print(pa_to_owner_driver_premium)

        # PA for Paid Driver
        a, pa_to_paid_driver_premium = str(premium1).split('PA Cover For Paid Driver (IMT-17)', 1)
        pa_to_paid_driver_premium, b = str(pa_to_paid_driver_premium).strip().split('\n', 1)
        print('*** PA Cover for Paid Driver')
        print(pa_to_paid_driver_premium)

        # Liblity for Paid Driver
        a, liability_to_paid_driver_premium = str(premium1).split('Legal Liability For Paid Driver (IMT-28)', 1)
        liability_to_paid_driver_premium, b = str(liability_to_paid_driver_premium).strip().split('\n', 1)
        print('*** liblity for Paid Driver')
        print(liability_to_paid_driver_premium)

        # Third Party Liability For Bi-Fuel Kit
        a, liability_bi_fuel_kit = str(premium1).split('Third Party Liability For Bi-Fuel Kit', 1)
        liability_bi_fuel_kit, b = str(liability_bi_fuel_kit).strip().split('\n', 1)
        print('*** liability fuel kit Driver')
        print(liability_bi_fuel_kit)

        # PA Cover For 5 Person for Rs 200000 (IMT-16)
        a, liability_5_person = str(premium1).split('PA Cover For 5 Person for Rs 200000 (IMT-16)', 1)
        liability_5_person, b = str(liability_5_person).strip().split('\n', 1)
        print('*** liability fuel kit Driver')
        print(liability_5_person)
        total_tax=0
        # CGST
        try:
            a, cgst = str(premium1).split('CGST ( 9 %)', 1)
            cgst, b = str(cgst).strip().split('\n', 1)
            print('*** CGST')
            print(cgst)
            total_tax=total_tax+int(str(cgst).strip())

        except ValueError:
            cgst=None

        #  sgst
        try:
            a, sgst = str(premium1).split('SGST ( 9 %)', 1)
            sgst, b = str(sgst).strip().split('\n', 1)
            print('*** SGST')
            print(sgst)
            total_tax=total_tax+int(str(sgst).strip())

        except ValueError:
            sgst = None

        #  Igst
        try:
            a, igst = str(premium1).split('IGST ( 18 %)', 1)
            igst, b = str(igst).strip().split('\n', 1)
            print('*** IGST')
            print(igst)
            total_tax=total_tax+int(str(igst).strip())

        except ValueError:
            igst=None

        #  ugst
        try:
            a, ugst = str(premium1).split('UGST ( 9 %)', 1)
            ugst, b = str(ugst).strip().split('\n', 1)
            print('*** UGST')
            print(ugst)
            total_tax=total_tax+int(str(ugst).strip())

        except ValueError:
            ugst=None

        # Total premium a+b
        a, total_premium = str(premium1).split('Total Premium (A \n + B\n)', 1)
        total_premium, b = str(total_premium).strip().split('\n', 1)
        print('*** Total Premium')
        print(total_premium)

        # Gross Premium
        a, gross_premium = str(premium1).split('Gross Premium (INR)', 1)
        gross_premium, b = str(gross_premium).strip().split('\n', 1)
        print('*** Gross Premium')
        print(gross_premium)
    else:
        basic_tp_premium =None
        pa_to_owner_driver_premium=None
        liability_to_paid_driver_premium=None
        pa_to_paid_driver_premium=None
        liability_bi_fuel_kit=None
        liability_5_person = None

        total_tax=0
        # CGST
        try:
            a, cgst = str(premium1).split('CGST ( 9 %)', 1)
            cgst, b = str(cgst).strip().split('\n', 1)
            print('*** CGST')
            print(cgst)
            total_tax=total_tax+int(str(cgst).strip())
        except ValueError:
            cgst=None

        #  sgst
        try:
            a, sgst = str(premium1).split('SGST ( 9 %)', 1)
            sgst, b = str(sgst).strip().split('\n', 1)
            print('*** SGST')
            print(sgst)
            total_tax=total_tax+int(str(sgst).strip())

        except ValueError:
            sgst = None

        #  Igst
        try:
            a, igst = str(premium1).split('IGST ( 18 %)', 1)
            igst, b = str(igst).strip().split('\n', 1)
            print('*** IGST')
            print(igst)
            total_tax=total_tax+int(str(igst).strip())

        except ValueError:
            igst=None

        #  ugst
        try:
            a, ugst = str(premium1).split('UGST ( 9 %)', 1)
            ugst, b = str(ugst).strip().split('\n', 1)
            print('*** UGST')
            print(ugst)
            total_tax=total_tax+int(str(ugst).strip())

        except ValueError:
            ugst=None
        print('TP premium not found')
        a, total_premium = str(premium1).split('Total Premium (A \n)', 1)
        total_premium, b = str(total_premium).strip().split('\n', 1)
        print('*** Total Premium')
        print(total_premium)

        # Gross Premium
        a, gross_premium = str(premium1).split('Gross Premium (INR)', 1)
        gross_premium, b = str(gross_premium).strip().split('\n', 1)
        print('*** Gross Premium')
        print(gross_premium)

    # Add data in excel and save
    data_Sheet_daily_workbook.cell(row=child_rowno + 2, column=12).value = pdf_policy_number
    data_Sheet_daily_workbook.cell(row=child_rowno + 2, column=13).value = vehical_reg_num
    data_Sheet_daily_workbook.cell(row=child_rowno + 2, column=14).value = fuel_type
    data_Sheet_daily_workbook.cell(row=child_rowno + 2, column=15).value = chassis_num
    data_Sheet_daily_workbook.cell(row=child_rowno + 2, column=16).value = engine_num
    data_Sheet_daily_workbook.cell(row=child_rowno + 2, column=17).value = yom
    data_Sheet_daily_workbook.cell(row=child_rowno + 2, column=18).value = dept_idv
    data_Sheet_daily_workbook.cell(row=child_rowno + 2, column=19).value = ele_acc_value
    data_Sheet_daily_workbook.cell(row=child_rowno + 2, column=20).value = non_elc_acc_value
    data_Sheet_daily_workbook.cell(row=child_rowno + 2, column=21).value = cng_lpg_value
    data_Sheet_daily_workbook.cell(row=child_rowno + 2, column=22).value = total_idv
    data_Sheet_daily_workbook.cell(row=child_rowno + 2, column=23).value = basic_premium
    data_Sheet_daily_workbook.cell(row=child_rowno + 2, column=24).value = ele_accessories
    data_Sheet_daily_workbook.cell(row=child_rowno + 2, column=25).value = non_ele_accessories

    data_Sheet_daily_workbook.cell(row=child_rowno + 2, column=26).value = lpg_cng_premium
    data_Sheet_daily_workbook.cell(row=child_rowno + 2, column=27).value = total_basic_premium
    data_Sheet_daily_workbook.cell(row=child_rowno + 2, column=28).value = ncb
    data_Sheet_daily_workbook.cell(row=child_rowno + 2, column=29).value = ncb_value
    data_Sheet_daily_workbook.cell(row=child_rowno + 2, column=30).value = other_discount
    data_Sheet_daily_workbook.cell(row=child_rowno + 2, column=31).value = total_od_premium
    data_Sheet_daily_workbook.cell(row=child_rowno + 2, column=32).value = net_od_premium
    data_Sheet_daily_workbook.cell(row=child_rowno + 2, column=33).value = str(add_on_text)
    data_Sheet_daily_workbook.cell(row=child_rowno + 2, column=34).value = addon_premium_final
    data_Sheet_daily_workbook.cell(row=child_rowno + 2, column=35).value = imt58_premium
    data_Sheet_daily_workbook.cell(row=child_rowno + 2, column=36).value = basic_tp_premium
    data_Sheet_daily_workbook.cell(row=child_rowno + 2, column=37).value = pa_to_owner_driver_premium
    data_Sheet_daily_workbook.cell(row=child_rowno + 2, column=38).value = pa_to_paid_driver_premium
    data_Sheet_daily_workbook.cell(row=child_rowno + 2, column=39).value = liability_bi_fuel_kit
    data_Sheet_daily_workbook.cell(row=child_rowno + 2, column=40).value = liability_5_person
    data_Sheet_daily_workbook.cell(row=child_rowno + 2, column=41).value = liability_to_paid_driver_premium
    data_Sheet_daily_workbook.cell(row=child_rowno + 2, column=42).value = total_premium
    data_Sheet_daily_workbook.cell(row=child_rowno + 2, column=43).value = cgst
    data_Sheet_daily_workbook.cell(row=child_rowno + 2, column=44).value = sgst
    data_Sheet_daily_workbook.cell(row=child_rowno + 2, column=45).value = igst
    data_Sheet_daily_workbook.cell(row=child_rowno + 2, column=46).value = ugst

    data_Sheet_daily_workbook.cell(row=child_rowno + 2, column=47).value = total_tax
    data_Sheet_daily_workbook.cell(row=child_rowno + 2, column=48).value = gross_premium

    daily_workbook.save(daily_file_path)
    daily_workbook.save(daily_file_path_backup)






